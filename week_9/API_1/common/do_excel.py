# -*- coding: utf-8 -*-
# @Time    : 2019/3/11 20:31
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py
from openpyxl import load_workbook

class DoExcel:
    '''该类完成测试数据的读取以及测试结果的写回'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#Excel工作簿文件名或地址
        self.sheet_name=sheet_name#表单名

    def read_data(self):
        '''从Excel读取数据，有返回值'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        #唯一的要求是什么？每一行数据要在一起 {} []
        #如何把每一行的数据存到一个空间里面去？ []
        #开始读取数据
        test_data=[]
        for i in range(2,sheet.max_row+1):
            row_data={}
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Module']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Url']=sheet.cell(i,4).value
            row_data['Method']=sheet.cell(i,5).value
            row_data['Params']=sheet.cell(i,6).value
            row_data['ExpectedResult']=sheet.cell(i,7).value
            test_data.append(row_data)
        wb.close()
        return test_data

    def write_back(self,row,col,value):
        '''写回测试结果到Excel中'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        sheet.cell(row,col).value=value

        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
    file_name='test_api.xlsx'
    sheet_name='test_case'
    test_data=DoExcel(file_name,sheet_name).read_data()
    print(test_data)
