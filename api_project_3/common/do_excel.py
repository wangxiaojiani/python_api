# -*- coding: utf-8 -*-
# Created by xj on 2019/3/10

from openpyxl import workbook,load_workbook
from api_project_3.common.api_conf import MyConfig
from api_project_3.common.project_path import test_data_path
import re



class DoExcel:
    def __init__(self,file_path,sheetname):
        self.file_path =file_path
        self.sheetname=sheetname
    def read_excel_tel(self):#定义一个读取注册电话的函数
        wb=load_workbook(self.file_path)
        sheetname=wb["Tel"] #在这里写死sheet表名
        tel_num =sheetname.cell(row=1,column=2).value
        return tel_num #返回电话号码
    def update_excel_tel(self,value): #定义一个写回注册电话的函数
        wb=load_workbook(self.file_path)
        sheetname=wb["Tel"]
        sheetname.cell(row=1,column=2).value=value
        wb.save(self.file_path)
        wb.close()
    def updata_excel_amount(self,amount):#定义一个写回leaveamount的函数
        wb=load_workbook(self.file_path)
        sheetname=wb['Tel']
        sheetname.cell(row=2,column=2).value=amount
        wb.save(self.file_path)
        wb.close()

    def read_excel(self,SECTION):
        button = MyConfig().get_list(SECTION,"button")  # 从配置文件中读取指定用例
        # mobile = MyConfig().get_list("RECHARGE", "m_mobile") #从配置文件中读取手机号码 类型为字符串

        wb = load_workbook(self.file_path) #打开excel
        sheetname = wb[self.sheetname] #定位表单
        test_data=[] #每行存储的测试数据列表
        new_tel = self.read_excel_tel()
        #方法1：通过列表的方式来存储每行数据
        # if button =="ALL":
        #     for i in range(2,sheetname.max_row+1):
        #         row_list=[] #每行数据存放在一个小列表中
        #         for j in range(1,8):
        #             row_list.append(sheetname.cell(row=i,column=j).value)
        #         test_data.append(row_list)
        # else:
        #     for i in button:
        #         row_list=[]
        #         for j in range(1,8):
        #             row_list.append(sheetname.cell(row=i+1,column=j).value)
        #         test_data.append(row_list)
        # return test_data

        #====方法2：通过字典的方式来存储每行数据
        # if button =="ALL":
        #     for i in range(2,sheetname.max_row+1):
        #         row_dict={} #每行数据存放在一个小列表中
        #         row_dict["Case_id"]=sheetname.cell(row=i,column=1).value
        #         row_dict["Moudle"] = sheetname.cell(row=i, column=2).value
        #         row_dict["Title"] = sheetname.cell(row=i, column=3).value
        #         row_dict["Method"] = sheetname.cell(row=i + 1, column=4).value
        #         row_dict["Url"] = sheetname.cell(row=i, column=5).value
        #         row_dict["Params"] = sheetname.cell(row=i, column=6).value
        #         row_dict["ExpectResult"] = sheetname.cell(row=i, column=7).value
        #         test_data.append(row_dict)
        # else:
        #     for i in button:
        #         row_dict ={}  # 每行数据存放在一个字典中
        #         row_dict["Case_id"] = sheetname.cell(row=i+1, column=1).value
        #         row_dict["Moudle"] = sheetname.cell(row=i+1, column=2).value
        #         row_dict["Title"] = sheetname.cell(row=i+1, column=3).value
        #         row_dict["Method"] = sheetname.cell(row=i + 1, column=4).value
        #         row_dict["Url"] = sheetname.cell(row=i+1, column=5).value
        #         row_dict["Params"] = sheetname.cell(row=i+1, column=6).value
        #         row_dict["ExpectResult"] = sheetname.cell(row=i+1, column=7).value
        #         test_data.append(row_dict)
        # return test_data

        #方法三 字典初始化2
        for i in range(2, sheetname.max_row + 1):
            row_dict={} #每行数据存放在一个小列表中
            row_dict["Case_id"]=sheetname.cell(row=i,column=1).value
            row_dict["Moudle"] = sheetname.cell(row=i, column=2).value
            row_dict["Title"] = sheetname.cell(row=i,column=3).value
            row_dict["Method"] = sheetname.cell(row=i,column=4).value
            row_dict["Url"] = sheetname.cell(row=i, column=5).value
            row_dict["Params"] = sheetname.cell(row=i,column=6).value
            row_dict["Sql"]=sheetname.cell(row=i,column=7).value
            row_dict["ExpectResult"] = sheetname.cell(row=i,column=8).value
            if sheetname.cell(row=i,column=6).value.find('tel') != -1:   #找到了tel就进行替换
                new_params=sheetname.cell(row=i,column=6).value.replace('tel',str(new_tel)) #repalce 两个参数必须是字符串
                self.update_excel_tel(new_tel+1)
            # elif re.search('#(.+?)#',sheetname.cell(row=i,column=6).value): #判定是否能从字符串中读取到m_mobile
            #     p ='#(.+?)#' #设定手机号的匹配模式
            #     new_params=re.sub(p,mobile,sheetname.cell(row=i,column=6).value) #将从excel中匹配到的字符串用 配置文件中对应的值进行替换
            else:
                new_params=sheetname.cell(row=i,column=6).value
            row_dict["Params"] = new_params
            test_data.append(row_dict)
        wb.close()
        final_data = []  # 空列表 存储最终的测试用例数据
        if button == 'all':  # 如果case_id==all 那就获取所有的用例数据
            final_data = test_data  # 把测试用例赋值给final_data这个变量
        else:  # 否则 如果是列表 那就获取列表里面指定id的用例的数据
            for i in button:  # 遍历case_id 里面的值
                final_data.append(test_data[i - 1])  # 对应关系要清晰
        return final_data

    def write_back(self,row,colunm,value):#向excel中写回测试结果
        wb=load_workbook(self.file_path)
        sheetname = wb[self.sheetname]
        sheetname.cell(row,colunm).value=value
        wb.save(self.file_path)
        wb.close()
    def creat_sheet(self):#创建表单
        wb=workbook.Workbook()
        wb.create_sheet(self.sheetname)
        wb.save(self.file_path)

if __name__ == '__main__':
    button =MyConfig().get_list("INVEST","button")
    print(button)
    test_data =DoExcel(test_data_path,"invest").read_excel("INVEST")
    print(test_data)
