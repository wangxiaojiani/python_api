# -*- coding: utf-8 -*-
# Created by xj on 2019/3/9

from openpyxl import workbook,load_workbook
from week_7.exercise.cof import MyConfig
from week_7.exercise.logf import MyLog
logger=MyLog()
button =MyConfig('conf.config').read_list('TESTCASE','button')
class DoExcel:
    def __init__(self,filename,sheetname):
        self.filename =filename
        self.sheetname=sheetname
    def read_excel(self,button):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]
        test_data=[]
        if button =='ALL':
            for i in range(2,sheet.max_row+1):
                row_list=[]
                for j in range(1,6):
                    row_list.append(sheet.cell(row=i,column=j).value)
                test_data.append(row_list)
            logger.info("结束执行用例")
        else:
            logger.info("开始执行用例")
            for i in button:
                row_list=[]
                for j in range(1,6):
                    row_list.append(sheet.cell(row=i+1,column=j).value)
                test_data.append(row_list)
            logger.info("用例执行完毕")
        return test_data
    def write_back(self,row,column,value):
        wb =load_workbook(self.filename)
        sheet=wb[self.sheetname]
        sheet.cell(row,column).value=value
        wb.save(self.filename)
        wb.close()
    def creat_sheet(self):
        wb=workbook.Workbook()
        wb.create_sheet(self.sheetname)
        wb.save(self.filename)
if __name__ == '__main__':
    wb =DoExcel('ppy14.xlsx','Sheet1')
    s=wb.read_excel(button)





