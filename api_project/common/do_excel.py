# -*- coding: utf-8 -*-
# Created by xj on 2019/3/10

from openpyxl import workbook,load_workbook
from common.api_conf import MyConfig

button =MyConfig().get_list("SELECT","button")#从配置文件中读取指定用例


class DoExcel:
    def __init__(self,filename,sheetname):
        self.filename =filename
        self.sheetname=sheetname
    def read_excel(self,button):
        wb = load_workbook(self.filename) #打开excel
        sheetname = wb[self.sheetname] #定位表单
        test_data=[] #每行存储的测试数据列表
        if button =="ALL":
            for i in range(2,sheetname.max_row+1):
                row_list=[] #每行数据存放在一个小列表中
                for j in range(1,8):
                    row_list.append(sheetname.cell(row=i,column=j).value)
                test_data.append(row_list)
        else:
            for i in button:
                row_list=[]
                for j in range(1,8):
                    row_list.append(sheetname.cell(row=i+1,column=j).value)
                test_data.append(row_list)
        return test_data
    def write_back(self,row,colunm,value):#向excel中写回测试结果
        wb=load_workbook(self.filename)
        sheetname = wb[self.sheetname]
        sheetname.cell(row,colunm).value=value
        wb.save(self.filename)
        wb.close()
    def creat_sheet(self):#创建表单
        wb=workbook.Workbook()
        wb.create_sheet(self.sheetname)
        wb.save(self.filename)

if __name__ == '__main__':
    T = DoExcel("API_TEST.xlsx","Sheet1").read_excel()
    print(T)

