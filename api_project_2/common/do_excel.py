# -*- coding: utf-8 -*-
# Created by xj on 2019/3/10

from openpyxl import workbook,load_workbook
from api_project_2.common.api_conf import MyConfig
from api_project_2.common.project_path import test_data_path

button =MyConfig().get_list("SELECT","button")#从配置文件中读取指定用例

class DoExcel:
    def __init__(self,filename,sheetname):
        self.filename =filename
        self.sheetname=sheetname
    def read_excel_tel(self):#定义一个读取注册电话的函数
        wb =load_workbook(self.filename,self.sheetname)
        sheetname=wb[self.sheetname]
        test_data_tel={}
        tel_key=sheetname.cell(row=1,column=1).value
        tel_num=sheetname.cell(row=1,column=2).value
        test_data_tel[tel_key]=tel_num
        return  test_data_tel #返回的是字典


    def read_excel(self,button):
        wb = load_workbook(self.filename) #打开excel
        sheetname = wb[self.sheetname] #定位表单
        test_data=[] #每行存储的测试数据列表
        #方法1：通过列表的方式来存储每行数据
        # if button =="ALL":
        #     for i in range(2,sheetname.max_row+1):
        #         row_list=[] #每行数据存放在一个小列表中
        #         for j in range(1,8):
        #             row_list.append(sheetname.cell(row=i,column=j).value)
        #         test_data.append(row_list)
        # else:
        #     for i in button:
        #         row_list=[]
        #         for j in range(1,8):
        #             row_list.append(sheetname.cell(row=i+1,column=j).value)
        #         test_data.append(row_list)
        # return test_data

        #====方法2：通过字典的方式来存储每行数据
        # if button =="ALL":
        #     for i in range(2,sheetname.max_row+1):
        #         row_dict={} #每行数据存放在一个小列表中
        #         row_dict["Case_id"]=sheetname.cell(row=i,column=1).value
        #         row_dict["Moudle"] = sheetname.cell(row=i, column=2).value
        #         row_dict["Title"] = sheetname.cell(row=i, column=3).value
        #         row_dict["Method"] = sheetname.cell(row=i + 1, column=4).value
        #         row_dict["Url"] = sheetname.cell(row=i, column=5).value
        #         row_dict["Params"] = sheetname.cell(row=i, column=6).value
        #         row_dict["ExpectResult"] = sheetname.cell(row=i, column=7).value
        #         test_data.append(row_dict)
        # else:
        #     for i in button:
        #         row_dict ={}  # 每行数据存放在一个字典中
        #         row_dict["Case_id"] = sheetname.cell(row=i+1, column=1).value
        #         row_dict["Moudle"] = sheetname.cell(row=i+1, column=2).value
        #         row_dict["Title"] = sheetname.cell(row=i+1, column=3).value
        #         row_dict["Method"] = sheetname.cell(row=i + 1, column=4).value
        #         row_dict["Url"] = sheetname.cell(row=i+1, column=5).value
        #         row_dict["Params"] = sheetname.cell(row=i+1, column=6).value
        #         row_dict["ExpectResult"] = sheetname.cell(row=i+1, column=7).value
        #         test_data.append(row_dict)
        # return test_data

        #方法三 字典初始化2
        for i in range(2, sheetname.max_row + 1):
            row_dict={} #每行数据存放在一个小列表中
            row_dict["Case_id"]=sheetname.cell(row=i,column=1).value
            row_dict["Moudle"] = sheetname.cell(row=i, column=2).value
            row_dict["Title"] = sheetname.cell(row=i, column=3).value
            row_dict["Method"] = sheetname.cell(row=i + 1, column=4).value
            row_dict["Url"] = sheetname.cell(row=i, column=5).value
            row_dict["Params"] = sheetname.cell(row=i, column=6).value
            row_dict["ExpectResult"] = sheetname.cell(row=i, column=7).value
            test_data.append(row_dict)
        wb.close()
        final_data = []  # 空列表 存储最终的测试用例数据
        if button == 'all':  # 如果case_id==all 那就获取所有的用例数据
            final_data = test_data  # 把测试用例赋值给final_data这个变量
        else:  # 否则 如果是列表 那就获取列表里面指定id的用例的数据
            for i in button:  # 遍历case_id 里面的值
                final_data.append(test_data[i - 1])  # 对应关系要清晰
        return final_data


    def write_back(self,row,colunm,value):#向excel中写回测试结果
        wb=load_workbook(self.filename)
        sheetname = wb[self.sheetname]
        sheetname.cell(row,colunm).value=value
        wb.save(self.filename)
        wb.close()
    def creat_sheet(self):#创建表单
        wb=workbook.Workbook()
        wb.create_sheet(self.sheetname)
        wb.save(self.filename)

if __name__ == '__main__':
    button = MyConfig().get_list("SELECT","button")
    T = DoExcel(test_data_path,"Sheet1").read_excel(button)
    print(len(T))
    print(T)

