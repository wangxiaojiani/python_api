# -*- coding: utf-8 -*-
# Created by xj on 2019/3/10

from openpyxl import workbook,load_workbook


class DoExcel:
    def __init__(self,filename,sheetname):
        self.filename =filename
        self.sheetname=sheetname
    def read_excel(self,start,end):
        wb = load_workbook(self.filename) #打开excel
        sheetname = wb[self.sheetname] #定位表单
        test_data=[] #每行存储的测试数据列表
        for i in range(start,end+1):
            row_list=[] #每行数据存放在一个小列表中
            for j in range(1,7):
                row_list.append(sheetname.cell(row=i,column=j).value)
            test_data.append(row_list)
        return test_data
    def write_back(self,row,colunm,value):#向excel中写回测试结果
        wb=load_workbook(self.filename)
        sheetname = wb[self.sheetname]
        sheetname.cell(row,colunm).value=value
        wb.save(self.filename)
        wb.close()
    def creat_sheet(self):#创建表单
        wb=workbook.Workbook()
        wb.create_sheet(self.sheetname)
        wb.save(self.filename)

if __name__ == '__main__':
    T = DoExcel("API_TEST.xlsx","Sheet1").read_excel()
    print(T)

