# -*- coding: utf-8 -*-
#@Time      :2019/9/30    22:53
#@Author    :xj
#@Email     :1027867874@qq.com
#@File      :do_excel.py
#@Software  :PyCharm

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from test_auto_interface.comon.api_conf import MyConfig
class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name
    def do_excel(self,SECTION):
        button=MyConfig().get_list(SECTION,'button')
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        mobile=self.get_init_data ()
        test_data=[]
        for i in range(2,sheet.max_row+1):
            sub_data={}
            sub_data['Case_id']=sheet.cell(i,1).value
            sub_data['Modle']=sheet.cell(i,2).value
            sub_data['Title'] = sheet.cell (i, 3).value
            sub_data['Method'] = sheet.cell (i, 4).value
            sub_data['Url'] = sheet.cell (i, 5).value
            sub_data['Params'] = sheet.cell (i, 6).value
            sub_data['ExpectResult']=sheet.cell(i,8).value
            if sheet.cell(i,5).value.find('${mobile}') !=-1:
                new_params =sheet.cell(i,5).value.replace('${mobile}',str(mobile))
            else:
                new_params=sheet.cell(i,5).value
                self.update_mobile_data(mobile+1)
            sub_data['Params']=new_params

            test_data.append(sub_data)
        final_data=[]
        if button=='all':
            final_data=test_data
        else:
            for i in button:
                final_data=test_data[i-1]
        return test_data
    def write_back(self,row,cloumn,value):
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        sheet.cell(row,cloumn).value=value
        wb.save(self.file_path)
    def creat_sheet(self):
        wb=Workbook()
        wb.create_sheet(self.sheet_name)
        wb.save(self.file_path)

    def get_init_data(self):
        """
        获取初始化手机号
        """
        wb=load_workbook(self.file_path)
        sheet=wb['Tel']
        mobile=sheet.cell(1,1).value
        return mobile #获取手机号码 注意格式是int

    def update_mobile_data(self,value):
        wb=load_workbook(self.file_path)
        sheet=wb['Tel']
        sheet.cell(1,1).value  =value
        wb.save(self.file_path)





if __name__=='__main__':




    wb = load_workbook (r'E:\learning_python_class\test_auto_interface\comon\nihao.xlsx')
    sheet = wb['1']
    for i in range(1,sheet.max_row+1):
        v=sheet.cell(i,1).value
        print(v)
        print(type(v))
        print(type(eval(v)))



